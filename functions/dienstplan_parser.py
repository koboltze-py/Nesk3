"""
Dienstplan-Parser für Excel-Dateien
Adaptiert aus Nesk2/Function/dienstplan_parser.py für Nesk3
Erkennt Namen, Dienstzeiten, Betreuer/Dispo-Unterscheidung dynamisch
"""
import os
import sys
import openpyxl
import re
from pathlib import Path
from datetime import datetime, time
from typing import Optional
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class DienstplanParser:
    """Parser für Dienstplan-Excel-Dateien.
    Sucht die Header-Zeile (NAME, DIENST, BEGINN, ENDE) automatisch
    innerhalb der ersten 20 Zeilen.
    """

    BETREUER_KATEGORIEN = ['T', 'T10', 'N', 'N10', 'NF', 'FB1', 'FB2', 'FB']
    DISPO_KATEGORIEN    = ['DT', 'DT3', 'DN', 'DN3', 'D']

    # Dienste die still (ohne Warnung) komplett ignoriert werden
    STILLE_DIENSTE: frozenset = frozenset({'R', 'B1', 'B2'})

    # Vollständig ausgeschlossene Personen (Vorname Nachname, lowercase, fix)
    AUSGESCHLOSSENE_VOLLNAMEN: frozenset = frozenset({'lars peters'})

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook   = None
        self.sheet      = None
        self.column_map = None
        self.unbekannte_dienste: set = set()

    # ------------------------------------------------------------------
    # Öffentliche Methode
    # ------------------------------------------------------------------

    def parse(self) -> dict:
        """
        Parst Excel-Datei und liefert Dienstplan-Daten.

        Returns:
            {
                'success': bool,
                'betreuer': list[dict],
                'dispo':    list[dict],
                'kranke':   list[dict],
                'error':    str | None,
                'unbekannte_dienste': list[str]
            }
        """
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            self.sheet    = self.workbook.active

            self.column_map = self._find_columns()
            if not self.column_map:
                return {
                    'success': False,
                    'betreuer': [], 'dispo': [], 'kranke': [],
                    'error': 'Header-Zeile nicht gefunden (benötigt: NAME, DIENST).',
                    'unbekannte_dienste': []
                }

            betreuer_liste = []
            dispo_liste    = []
            kranke_liste   = []
            alle_nachnamen = []

            for row in self.sheet.iter_rows(min_row=1, values_only=False):
                person = self._parse_row(row)
                if person:
                    alle_nachnamen.append(person['nachname'])
                    if person['ist_krank']:
                        kranke_liste.append(person)
                    elif person['ist_dispo']:
                        dispo_liste.append(person)
                    else:
                        betreuer_liste.append(person)

            # Ausgeschlossene Personen herausfiltern
            try:
                from functions.settings_functions import get_ausgeschlossene_namen
                settings_ausgeschlossen = set(get_ausgeschlossene_namen())
            except Exception:
                settings_ausgeschlossen = set()

            alle_ausgeschlossen = self.AUSGESCHLOSSENE_VOLLNAMEN | settings_ausgeschlossen

            def _filter_ausgeschlossen(lst):
                return [p for p in lst
                        if p['vollname'].lower() not in alle_ausgeschlossen]

            betreuer_liste = _filter_ausgeschlossen(betreuer_liste)
            dispo_liste    = _filter_ausgeschlossen(dispo_liste)
            kranke_liste   = _filter_ausgeschlossen(kranke_liste)

            # Doppelte Nachnamen → Initial anhängen
            nachname_counts  = Counter(alle_nachnamen)
            doppelte         = {n for n, c in nachname_counts.items() if c > 1}
            for gruppe in (betreuer_liste, dispo_liste, kranke_liste):
                self._generate_display_names(gruppe, doppelte)

            return {
                'success': True,
                'betreuer': betreuer_liste,
                'dispo':    dispo_liste,
                'kranke':   kranke_liste,
                'error':    None,
                'unbekannte_dienste': list(self.unbekannte_dienste)
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'betreuer': [], 'dispo': [], 'kranke': [],
                'error': str(e),
                'unbekannte_dienste': []
            }
        finally:
            if self.workbook:
                self.workbook.close()

    # ------------------------------------------------------------------
    # Interne Hilfsmethoden
    # ------------------------------------------------------------------

    def _find_columns(self) -> Optional[dict]:
        """Sucht Header-Zeile und gibt Spalten-Indizes zurück."""
        for row_idx in range(1, min(20, self.sheet.max_row + 1)):
            row        = list(self.sheet[row_idx])
            row_values = [cell.value for cell in row]

            name_idx = dienst_idx = beginn_idx = ende_idx = None

            for col_idx, value in enumerate(row_values):
                if value and isinstance(value, str):
                    v = value.strip().upper()
                    if v == 'NAME':
                        name_idx = col_idx
                    elif v == 'DIENST':
                        dienst_idx = col_idx
                    elif v == 'BEGINN':
                        beginn_idx = col_idx
                    elif v == 'ENDE':
                        ende_idx = col_idx

            if name_idx is not None and dienst_idx is not None:
                return {
                    'name':       name_idx,
                    'dienst':     dienst_idx,
                    'beginn':     beginn_idx,
                    'ende':       ende_idx,
                    'header_row': row_idx
                }
        return None

    def _parse_row(self, row) -> Optional[dict]:
        """Parst eine Zeile und gibt Person-Dict oder None zurück."""
        row_list = list(row)
        cells    = [cell.value if hasattr(cell, 'value') else cell for cell in row_list]

        max_col = max(
            self.column_map['name'],
            self.column_map['dienst'],
            self.column_map.get('beginn') or 0,
            self.column_map.get('ende')   or 0,
        )
        if len(cells) <= max_col:
            return None

        # Name prüfen
        name_cell = cells[self.column_map['name']]
        if not name_cell or not isinstance(name_cell, str):
            return None

        parsed_name = self._parse_name(name_cell)
        if not parsed_name:
            return None

        full_name = f"{parsed_name['vorname']} {parsed_name['nachname']}"

        # Zellfarbe prüfen (Bulmorfahrer = gelb)
        name_cell_obj   = row_list[self.column_map['name']]
        dienst_cell_obj = row_list[self.column_map['dienst']]
        ist_bulmorfahrer, zeilen_farbe, dienst_farbe, dienst_farbe_hex = \
            self._check_cell_colors(name_cell_obj, dienst_cell_obj)

        # Dienst-Kürzel
        dienst_text     = None
        dienst_kategorie = None
        ist_krank        = False

        raw_dienst = cells[self.column_map['dienst']]
        if raw_dienst:
            dienst_text = str(raw_dienst).strip().upper()
            if dienst_text in self.STILLE_DIENSTE:
                return None   # still ignorieren, keine Warnung
            if dienst_text in ('KRANK', 'K'):
                ist_krank = True
            elif dienst_text in self.BETREUER_KATEGORIEN or dienst_text in self.DISPO_KATEGORIEN:
                dienst_kategorie = dienst_text
            elif dienst_text:
                dienst_kategorie = dienst_text
                self.unbekannte_dienste.add(dienst_text)

        # Zeiten
        round_times = dienst_kategorie in self.DISPO_KATEGORIEN if dienst_kategorie else False
        start_zeit  = None
        end_zeit    = None

        if self.column_map.get('beginn') is not None and len(cells) > self.column_map['beginn']:
            start_zeit = self._parse_time(cells[self.column_map['beginn']], round_to_hour=round_times)
        if self.column_map.get('ende') is not None and len(cells) > self.column_map['ende']:
            end_zeit = self._parse_time(cells[self.column_map['ende']], round_to_hour=round_times)

        schicht_typ = self._ermittle_schichttyp(start_zeit, end_zeit)

        return {
            'vorname':          parsed_name['vorname'],
            'nachname':         parsed_name['nachname'],
            'vollname':         full_name,
            'anzeigename':      parsed_name['nachname'],
            'dienst_kategorie': dienst_kategorie,
            'start_zeit':       start_zeit,
            'end_zeit':         end_zeit,
            'schicht_typ':      schicht_typ,
            'ist_dispo':        dienst_kategorie in self.DISPO_KATEGORIEN if dienst_kategorie else False,
            'ist_krank':        ist_krank,
            'ist_bulmorfahrer': ist_bulmorfahrer,
            'zeilen_farbe':     zeilen_farbe,
            'dienst_farbe':     dienst_farbe,
            'dienst_farbe_hex': dienst_farbe_hex,
        }

    def _check_cell_colors(self, name_cell_obj, dienst_cell_obj):
        """Liest Zellfarben aus und erkennt Bulmorfahrer (gelb) und Zebra-Zeilen (grau)."""
        ist_bulmorfahrer = False
        zeilen_farbe     = None
        dienst_farbe     = None
        dienst_farbe_hex = None

        for cell_obj in (name_cell_obj, dienst_cell_obj):
            if not hasattr(cell_obj, 'fill') or not cell_obj.fill:
                continue
            try:
                fill = cell_obj.fill
                if fill.patternType in ('solid', 'solidFill'):
                    fg = fill.fgColor
                    if hasattr(fg, 'rgb') and fg.rgb:
                        color_hex = str(fg.rgb).upper()

                        if cell_obj == dienst_cell_obj and color_hex not in ('00000000', 'FFFFFFFF'):
                            dienst_farbe_hex = color_hex

                        if color_hex == 'FFF5F5F5':
                            zeilen_farbe = 'gray'

                        # Gelb erkennen (AARRGGBB oder RRGGBB)
                        if len(color_hex) >= 8:
                            rr_gg = color_hex[2:6]
                            bb    = color_hex[6:8]
                        elif len(color_hex) == 6:
                            rr_gg = color_hex[0:4]
                            bb    = color_hex[4:6]
                        else:
                            continue

                        if rr_gg == 'FFFF' and int(bb, 16) <= 0x4F:
                            ist_bulmorfahrer = True
                            dienst_farbe     = 'yellow'
                            break
            except Exception:
                pass

        return ist_bulmorfahrer, zeilen_farbe, dienst_farbe, dienst_farbe_hex

    def _parse_name(self, name_text: str) -> Optional[dict]:
        """Parst 'Nachname, Vorname' oder 'Vorname Nachname'."""
        name_text = name_text.strip()
        name_text = re.sub(r'[^\w\säöüÄÖÜß\-,]', '', name_text)

        if ',' in name_text:
            parts = name_text.split(',')
            if len(parts) == 2:
                nachname      = parts[0].strip()
                vorname_teil  = parts[1].strip()
                if not vorname_teil:
                    return None
                vorname = vorname_teil.split()[0] if vorname_teil.split() else None
                if not vorname:
                    return None
                return {'vorname': vorname, 'nachname': nachname}

        parts = name_text.split()
        if len(parts) >= 2:
            return {'vorname': parts[0], 'nachname': ' '.join(parts[1:])}

        return None

    def _parse_time(self, value, round_to_hour=False) -> Optional[str]:
        """Parst Zeitwert und gibt 'HH:MM'-String oder None zurück."""
        try:
            if isinstance(value, datetime):
                t = value.time()
            elif isinstance(value, time):
                t = value
            elif isinstance(value, str):
                value = value.strip()
                m = re.match(r'(\d{1,2}):(\d{2})', value)
                if m:
                    t = time(int(m.group(1)), int(m.group(2)))
                else:
                    m = re.match(r'(\d{2})(\d{2})', value)
                    if m:
                        t = time(int(m.group(1)), int(m.group(2)))
                    else:
                        return None
            else:
                return None

            if round_to_hour:
                t = time(t.hour, 0)
            return f"{t.hour:02d}:{t.minute:02d}"
        except Exception:
            return None

    def _ermittle_schichttyp(self, start_zeit: Optional[str], end_zeit: Optional[str]) -> Optional[str]:
        """Kategorisiert Schicht anhand der Startzeit."""
        if not start_zeit:
            return None
        try:
            hour = int(start_zeit.split(':')[0])
        except Exception:
            return None

        if 5 <= hour < 9 or 9 <= hour < 12:
            return 'tagdienst_vormittag'
        elif 12 <= hour < 15 or 15 <= hour < 19:
            return 'tagdienst_nachmittag'
        elif 19 <= hour < 23 or 23 <= hour <= 23:
            return 'nachtschicht_frueh'
        elif 0 <= hour < 5:
            return 'nachtschicht_spaet'
        return None

    def _generate_display_names(self, personen_liste: list, doppelte_nachnamen: set):
        """Hängt bei doppelten Nachnamen die ersten zwei Vorname-Buchstaben an."""
        for person in personen_liste:
            nachname = person['nachname']
            vorname  = person['vorname']
            if nachname in doppelte_nachnamen or '-' in nachname:
                if len(vorname) >= 2:
                    kurz = vorname[0].upper() + vorname[1].lower()
                elif len(vorname) == 1:
                    kurz = vorname[0].upper()
                else:
                    kurz = ''
                person['anzeigename'] = f"{nachname} {kurz}".strip()
            else:
                person['anzeigename'] = nachname
